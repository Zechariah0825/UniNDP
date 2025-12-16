import argparse
import csv
import os
import sys
from math import inf

import numpy as np
import tqdm

try:
    from openpyxl import Workbook
except ImportError as exc:  # pragma: no cover - runtime dependency
    raise SystemExit(
        "openpyxl is required for compile_e2e_pax.py. "
        "Please install it with `pip install openpyxl`."
    ) from exc

# Ensure project root (one level up from this script) is on sys.path so that
# `frontend`, `midend`, `backend`, etc. can be imported when running this file
# from the `script/` subdirectory.
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if PROJECT_ROOT not in sys.path:
    sys.path.insert(0, PROJECT_ROOT)

from frontend import *  # type: ignore
from midend import *  # type: ignore
from backend import *  # type: ignore
from sim import sim  # type: ignore
from tools import *  # type: ignore


def _normalize_model_name(name: str) -> str:
    """Convert 'Qwen3-8B' -> 'qwen3_8B' for filenames."""
    if "-" in name:
        family, size = name.split("-", 1)
        return f"{family.lower()}_{size}"
    return name.lower()


def _setup_architecture(architecture: str, workload: str = "mm"):
    """Configure SimConfig and return the Codegen class."""
    if architecture == "aim":
        SimConfig.read_from_yaml("./config/gddr6-aim.yaml")
        if workload == "mm":
            SimConfig.de_pu = [16]
        else:
            SimConfig.de_pu = [4]
        Codegen = aim16
    elif architecture == "aim8":
        SimConfig.read_from_yaml("./config/gddr6-aim.yaml")
        if workload == "mm":
            SimConfig.de_pu = [8]
        else:
            SimConfig.de_pu = [4]
        Codegen = aim8
    elif architecture == "hbm-pim":
        SimConfig.read_from_yaml("./config/hbm-pim.yaml")
        Codegen = hbmpim
    elif architecture == "lpddr-pax":
        SimConfig.read_from_yaml("./config/lpddr-pax.yaml")
        Codegen = hbmpim
    elif architecture == "upmem":
        SimConfig.read_from_yaml("./config/upmem.yaml")
        Codegen = upmem
    elif architecture == "dimmining":
        SimConfig.read_from_yaml("./config/dimmining.yaml")
        Codegen = dimmining
    else:
        raise ValueError(f"Unknown architecture: {architecture}")

    # Set PU level
    if architecture == "dimmining":
        SimConfig.pu_level = LEVEL.RA
    else:
        SimConfig.pu_level = LEVEL.DE

    return Codegen


def _build_design_space(mm_size, architecture: str, po2: bool, allow_under_ultize: bool):
    """Construct design space for a single MM workload."""
    partition_tool = Partition(require_power_of_2=po2)
    partition_space = partition_tool.get_partition_space_mm(mm_size)
    filtered_partition_space = partition_tool.choose_from_partition_space_mm(partition_space)
    if not allow_under_ultize:
        partition_space = filtered_partition_space

    design_space = []
    for compute_level, pu_num, partition in partition_space:
        simd_k, mkl_Input_to_row, simd_l, ml_Out_to_row = partition_tool.mem_partition_mm(
            mm_size, partition
        )
        for input_choice in reversed(mkl_Input_to_row):
            if architecture in ["aim", "aim8"]:
                if ml_Out_to_row:
                    output_choice = ml_Out_to_row[0]
                    design_space.append(
                        (compute_level, pu_num, partition, simd_k, input_choice, simd_l, output_choice)
                    )
                continue
            for output_choice in reversed(ml_Out_to_row):
                design_space.append(
                    (compute_level, pu_num, partition, simd_k, input_choice, simd_l, output_choice)
                )
    return design_space


def _select_baseline(design_space, architecture: str, mm_size):
    """Select baseline design point according to the same rules as compile.py."""
    baseline = None
    if architecture in ["aim", "aim8"]:
        for compute_level, pu_num, partition, simd_k, mkl_Input_to_row, simd_l, ml_Out_to_row in design_space:
            if partition[3][0] * partition[3][1] == 1 and (
                mkl_Input_to_row[0][0] * mkl_Input_to_row[0][2] * mkl_Input_to_row[0][3] == 1
                or mkl_Input_to_row[0][1] * simd_k == mm_size[1]
            ):
                baseline = (
                    compute_level,
                    pu_num,
                    partition,
                    simd_k,
                    mkl_Input_to_row,
                    simd_l,
                    ml_Out_to_row,
                )
                break
    elif architecture == "hbm-pim":
        for compute_level, pu_num, partition, simd_k, mkl_Input_to_row, simd_l, ml_Out_to_row in design_space:
            if partition[3][0] * partition[3][1] == 1 and mkl_Input_to_row[0][1] == 8:
                baseline = (
                    compute_level,
                    pu_num,
                    partition,
                    simd_k,
                    mkl_Input_to_row,
                    simd_l,
                    ml_Out_to_row,
                )
                break
    elif architecture == "upmem":
        for compute_level, pu_num, partition, simd_k, mkl_Input_to_row, simd_l, ml_Out_to_row in design_space:
            if partition[3][0] * partition[3][1] == 1 and (
                mkl_Input_to_row[0][0] * mkl_Input_to_row[0][2] * mkl_Input_to_row[0][3] == 1
                or mkl_Input_to_row[0][1] * simd_k == mm_size[1]
            ):
                baseline = (
                    compute_level,
                    pu_num,
                    partition,
                    simd_k,
                    mkl_Input_to_row,
                    simd_l,
                    ml_Out_to_row,
                )
                break
    elif architecture == "dimmining":
        for compute_level, pu_num, partition, simd_k, mkl_Input_to_row, simd_l, ml_Out_to_row in design_space:
            if partition[2][0] * partition[2][1] == 1 and (
                mkl_Input_to_row[0][0] * mkl_Input_to_row[0][2] * mkl_Input_to_row[0][3] == 1
                or mkl_Input_to_row[0][1] * simd_k == mm_size[1]
            ):
                baseline = (
                    compute_level,
                    pu_num,
                    partition,
                    simd_k,
                    mkl_Input_to_row,
                    simd_l,
                    ml_Out_to_row,
                )
                break

    if baseline is None and design_space:
        baseline = design_space[0]

    return baseline


def _run_codegen_and_sim(
    design_point,
    Codegen,
    po2: bool,
    cmd_threshold: float,
    run_sim: bool,
):
    """Helper: mapping + dram mapping + codegen (+ optional sim)."""
    (
        compute_level,
        pu_num,
        partition,
        simd_k,
        mkl_Input_to_row,
        simd_l,
        ml_Out_to_row,
    ) = design_point

    mapping_tool = Mapping(require_power_of_2=po2)
    hw_id_list = mapping_tool.assign_hw(partition)
    input_bank, input_row_offset, weight_bank, weight_row_offset, output_bank, output_row_offset = (
        mapping_tool.assign_dram(pu_num, mkl_Input_to_row, ml_Out_to_row, partition)
    )

    codegen_tool = Codegen(require_power_of_2=po2)
    if run_sim:
        codegen_tool.set_gen()
    gen_code, inst_count, predict_result = codegen_tool.codegen(
        "mm",
        compute_level,
        pu_num,
        partition,
        simd_k,
        mkl_Input_to_row,
        simd_l,
        ml_Out_to_row,
        hw_id_list,
        (input_bank, input_row_offset, weight_bank, weight_row_offset, output_bank, output_row_offset),
        cmd_threshold=cmd_threshold,
    )
    if gen_code is None:
        return None, predict_result, inst_count

    if run_sim:
        sim_result = sim(gen_code, silent=True, filename=None)
    else:
        sim_result = None

    return sim_result, predict_result, inst_count


def run_single_mm(
    M: int,
    K: int,
    N: int,
    B: int,
    architecture: str,
    po2: bool = True,
    quicksearch: bool = True,
    topk: int = 30,
    cmdthre: float = 3.0,
    allow_under_ultize: bool = False,
):
    """
    Compile a single MM workload and return
    (baseline_latency, baseline_inst_count, best_latency, best_inst_count).

    MM workload size is [M, K, N, B] following compile.py semantics.
    """
    mm_size = (M * B, K, N, 1)

    # 1) Setup arch and design space
    Codegen = _setup_architecture(architecture, workload="mm")
    design_space = _build_design_space(mm_size, architecture, po2, allow_under_ultize)

    if not design_space:
        raise RuntimeError("Empty design space for workload size {}".format(mm_size))

    # 2) Baseline
    baseline_dp = _select_baseline(design_space, architecture, mm_size)
    baseline_sim_result, _, baseline_inst_count = _run_codegen_and_sim(
        baseline_dp, Codegen, po2, cmd_threshold=0, run_sim=True
    )

    # 3) Search for best design
    if quicksearch:
        predict_result_list = []
        min_codelen = 0
        for design_point in design_space:
            thre = min_codelen * cmdthre
            sim_result, predict_result, inst_count = _run_codegen_and_sim(
                design_point, Codegen, po2, cmd_threshold=thre, run_sim=False
            )
            # Only keep design points where codegen succeeded
            if sim_result is None:
                # sim_result is None because we skipped sim; codegen success is judged by inst_count not None
                if inst_count is None:
                    continue
            predict_result_list.append((predict_result, design_point))
            if len(predict_result_list) > topk:
                predict_result_list = sorted(predict_result_list, key=lambda x: x[0])[:topk]

        candidate_designs = [x[1] for x in predict_result_list]
    else:
        candidate_designs = design_space

    best_result = inf
    best_inst_count = None
    for design_point in candidate_designs:
        sim_result, _, inst_count = _run_codegen_and_sim(
            design_point, Codegen, po2, cmd_threshold=0, run_sim=True
        )
        if sim_result is not None and sim_result < best_result:
            best_result = sim_result
            best_inst_count = inst_count

    return baseline_sim_result, baseline_inst_count, best_result, best_inst_count


def decode_mm_shapes(
    hdim: int,
    nheads: int,
    dhead: int,
    ff_scale: float,
    gqaheads: int,
    L: int,
    batch: int,
):
    """
    Return MM workload shapes for decode phase for the six ops:
    qkv_gen, qk, kv, out_proj, up, down.
    Shapes are tuples (M, K, N, B).
    """
    ff_dim = int(round(hdim * ff_scale))
    use_gqa = gqaheads is not None and gqaheads > 0 and gqaheads < nheads
    kv_heads = gqaheads if use_gqa else nheads
    groups = nheads // kv_heads if use_gqa else 1

    shapes = {}

    # Q/K/V share input (hdim) but produce three concatenated projections -> single larger MM
    shapes["qkv_gen"] = (1, hdim, 3 * hdim, batch)

    if use_gqa:
        group_dim = dhead * kv_heads  # = d_model / groups
        # Q: (groups, d_model/groups), K: (d_model/groups, L)
        shapes["qk"] = (groups, group_dim, L, batch)
        # S: (groups, nheads*L), V: (nheads*L, dhead)
        shapes["kv"] = (groups, nheads * L, dhead, batch)
    else:
        # MHA flattening
        # Q: (1, d_model), K: (d_model, L)
        shapes["qk"] = (1, hdim, L, batch)
        # S: (1, nheads*L), V: (nheads*L, dhead)
        shapes["kv"] = (1, nheads * L, dhead, batch)

    # Out projection and FFN
    shapes["out_proj"] = (1, hdim, hdim, batch)
    shapes["up"] = (1, hdim, ff_dim, batch)
    shapes["down"] = (1, ff_dim, hdim, batch)

    return shapes


def parse_models_csv(models_csv_path: str, target_model: str | None = None):
    """Yield model configs from models_llm.csv (optionally filtered by name)."""
    with open(models_csv_path, newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            name = row["name"]
            if target_model is not None and name != target_model:
                continue
            ndec = int(row["ndec"])
            hdim = int(row["hdim"])
            nheads = int(row["nheads"])
            dhead = int(row["dhead"])
            ff_scale = float(row["ff_scale"])
            gqaheads = int(row["gqaheads"])
            tklen_input = int(row.get("tkleninput", row.get("tklen_input", 32)))
            tklen_output = int(row.get("tklenoutput", row.get("tklen_output", tklen_input)))
            yield {
                "name": name,
                "ndec": ndec,
                "hdim": hdim,
                "nheads": nheads,
                "dhead": dhead,
                "ff_scale": ff_scale,
                "gqaheads": gqaheads,
                "tklen_input": tklen_input,
                "tklen_output": tklen_output,
            }


def main():
    parser = argparse.ArgumentParser(
        description="Generate per-operator decode speedup table over context length."
    )
    parser.add_argument(
        "--models-csv",
        type=str,
        default="script/models_llm.csv",
        help="CSV listing LLM models and hyper-parameters (including tkleninput/tklenoutput).",
    )
    parser.add_argument(
        "--model",
        type=str,
        default=None,
        help="If specified, only process this model name (must match 'name' column).",
    )
    parser.add_argument(
        "--batchsize",
        type=int,
        default=1,
        help="Batch size B, default is 1.",
    )
    parser.add_argument(
        "--architecture",
        "-A",
        type=str,
        default="aim",
        choices=["aim", "aim8", "hbm-pim", "upmem", "dimmining", "lpddr-pax"],
        help="Target architecture for compilation.",
    )
    parser.add_argument(
        "--po2",
        action="store_true",
        help="Require partition and mapping to be power of 2 (same as compile.py).",
    )
    parser.add_argument(
        "--cmdthre",
        "-T",
        type=float,
        default=3.0,
        help="Command length threshold factor for quicksearch.",
    )
    parser.add_argument(
        "--topk",
        "-K",
        type=int,
        default=30,
        help="Top-K design points kept after predictor-aided first round.",
    )
    parser.add_argument(
        "--allow-under-ultize",
        "-UU",
        action="store_true",
        help="Allow under-utilized design points in design space.",
    )
    parser.add_argument(
        "--output-dir",
        type=str,
        default="pruning_and_breakdown/e2e_pax",
        help="Directory to place per-model decode speedup CSVs.",
    )
    parser.add_argument(
        "--dump-workload",
        action="store_true",
        help="If set, also emit a workload CSV (Context_len/op/M/K/N/B) into --output-dir.",
    )

    args = parser.parse_args()

    os.makedirs(args.output_dir, exist_ok=True)

    # Prepare predictor and inst info for latency breakdown
    CodegenHeader = _setup_architecture(args.architecture, workload="mm")
    _codegen_tmp = CodegenHeader(require_power_of_2=args.po2)
    predictor_vec = np.array(_codegen_tmp.predictor, dtype=float)
    inst_len = len(_codegen_tmp.inst_info)

    # Column names for per-instruction latency sheets (must match inst_info order)
    inst_headers = [
        "device_pu",
        "device_pu_col",
        "device_pu_row_change",
        "device_reg2buf",
        "device_buf2reg",
        "device_buf2bk",
        "device_buf2bk_col",
        "device_bk2buf",
        "device_bk2buf_col",
        "device_bk2gb",
        "device_bk2gb_col",
        "device_gb2bk",
        "device_gb2bk_col",
        "host_read",
        "host_read_col",
        "host_write",
        "host_write_col",
        "host_write_device_buffer",
        "host_write_device_buffer_col",
        "host_write_pu_inbuf",
        "host_write_pu_inbuf_col",
        "host_read_mac_reg",
        "host_write_mac_reg",
    ]
    assert inst_len == len(
        inst_headers
    ), f"Mismatch between inst_info length ({inst_len}) and header ({len(inst_headers)})"

    # Batch size is processed here because it is used in decode_mm_shapes.
    batch=args.batchsize
    print(f"Processing batch size {batch}")
    # Iterate over models
    batch=args.batchsize
    for model_cfg in parse_models_csv(args.models_csv, args.model):
        name = model_cfg["name"]
        hdim = model_cfg["hdim"]
        nheads = model_cfg["nheads"]
        dhead = model_cfg["dhead"]
        ff_scale = model_cfg["ff_scale"]
        gqaheads = model_cfg["gqaheads"]
        tk_in = model_cfg["tklen_input"]
        tk_out = model_cfg["tklen_output"]

        print(f"Processing model {name} on architecture {args.architecture} (L={tk_in}..{tk_out}, batch={batch})")

        # 1) Compute shapes at initial L for all ops
        shapes_initial = decode_mm_shapes(hdim, nheads, dhead, ff_scale, gqaheads, tk_in, batch)

        # 2) Run single-mm compile for ops whose shapes don't change with L:
        #    qkv_gen, out_proj, up, down
        const_ops = ["qkv_gen", "out_proj", "up", "down"]
        const_baseline = {}
        const_best = {}
        const_speedup = {}
        const_best_inst = {}
        for op in const_ops:
            M, K, N, B = shapes_initial[op]
            print(f"  [const op] {op}: MM=({M},{K},{N},{B})")
            baseline_lat, baseline_ic, best_lat, best_ic = run_single_mm(
                M,
                K,
                N,
                B,
                architecture=args.architecture,
                po2=args.po2,
                quicksearch=True,
                topk=args.topk,
                cmdthre=args.cmdthre,
                allow_under_ultize=args.allow_under_ultize,
            )
            const_baseline[op] = baseline_lat
            const_best[op] = best_lat
            const_speedup[op] = baseline_lat / best_lat if best_lat and best_lat > 0 else 1.0
            const_best_inst[op] = best_ic

        # 3) Sweep context length for qk / kv
        rows_speedup = []
        rows_best = []
        rows_baseline = []
        # Per-op per-L instruction-level latency breakdown
        per_op_inst_rows = {op: [] for op in ["qkv_gen", "qk", "kv", "out_proj", "up", "down"]}
        workload_rows = [] if args.dump_workload else None

        for L in range(tk_in, tk_out + 1):
            shapes = decode_mm_shapes(hdim, nheads, dhead, ff_scale, gqaheads, L, batch)
            row_speed = {"Context_len": L}
            row_best = {"Context_len": L}
            row_base = {"Context_len": L}

            if workload_rows is not None:
                for op in ["qkv_gen", "qk", "kv", "out_proj", "up", "down"]:
                    M, K, N, B = shapes[op]
                    workload_rows.append([L, op, M, K, N, B])

            # qkv_gen / out_proj / up / down: reuse const results
            for op in const_ops:
                row_speed[op] = const_speedup[op]
                row_best[op] = const_best[op]
                row_base[op] = const_baseline[op]

            # qk / kv: recompute per L
            for op in ["qk", "kv"]:
                M, K, N, B = shapes[op]
                print(f"  [L={L}, B={batch}] {op}: MM=({M},{K},{N},{B})")
                baseline_lat, _, best_lat, best_ic = run_single_mm(
                    M,
                    K,
                    N,
                    B,
                    architecture=args.architecture,
                    po2=args.po2,
                    quicksearch=True,
                    topk=args.topk,
                    cmdthre=args.cmdthre,
                    allow_under_ultize=args.allow_under_ultize,
                )
                row_best[op] = best_lat
                row_base[op] = baseline_lat
                row_speed[op] = baseline_lat / best_lat if best_lat and best_lat > 0 else 1.0

                # store best inst counts per L for qk/kv
                per_op_inst_rows[op].append(
                    [L] + list((np.array(best_ic, dtype=float)[:inst_len] * predictor_vec).tolist())
                )

            # For const ops, their instruction latency contributions do not depend on L;
            # still record one row per L for completeness.
            for op in const_ops:
                best_ic = const_best_inst[op]
                per_op_inst_rows[op].append(
                    [L] + list((np.array(best_ic, dtype=float)[:inst_len] * predictor_vec).tolist())
                )

            rows_speedup.append(row_speed)
            rows_best.append(row_best)
            rows_baseline.append(row_base)

        # 4) Write per-model Excel workbook with three sheets
        norm_name = _normalize_model_name(name)
        out_path_xlsx = os.path.join(
            args.output_dir, f"{norm_name}_{args.architecture}_decode_metrics.xlsx"
        )

        wb = Workbook()

        header = ["Context_len", "qkv_gen", "qk", "kv", "out_proj", "up", "down"]

        # Sheet 1: speedup
        ws_speed = wb.active
        ws_speed.title = "speedup"
        ws_speed.append(header)
        for row in rows_speedup:
            ws_speed.append(
                [
                    row["Context_len"],
                    row["qkv_gen"],
                    row["qk"],
                    row["kv"],
                    row["out_proj"],
                    row["up"],
                    row["down"],
                ]
            )

        # Sheet 2: best latency
        ws_best = wb.create_sheet("best_latency")
        ws_best.append(header)
        for row in rows_best:
            ws_best.append(
                [
                    row["Context_len"],
                    row["qkv_gen"],
                    row["qk"],
                    row["kv"],
                    row["out_proj"],
                    row["up"],
                    row["down"],
                ]
            )

        # Sheet 3: baseline latency
        ws_base = wb.create_sheet("baseline_latency")
        ws_base.append(header)
        for row in rows_baseline:
            ws_base.append(
                [
                    row["Context_len"],
                    row["qkv_gen"],
                    row["qk"],
                    row["kv"],
                    row["out_proj"],
                    row["up"],
                    row["down"],
                ]
            )

        # Additional sheets: per-operator instruction-level latency breakdown
        inst_header_row = ["Context_len"] + inst_headers
        for op in ["qkv_gen", "qk", "kv", "out_proj", "up", "down"]:
            ws_op = wb.create_sheet(op)
            ws_op.append(inst_header_row)
            for row in per_op_inst_rows[op]:
                ws_op.append(row)

        os.makedirs(args.output_dir, exist_ok=True)
        wb.save(out_path_xlsx)

        if workload_rows is not None:
            workload_name = f"{norm_name}_{args.architecture}_decode_workload.csv"
            workload_path = os.path.join(args.output_dir, workload_name)
            with open(workload_path, "w", newline="") as wf:
                writer = csv.writer(wf)
                writer.writerow(["Context_len", "op", "M", "K", "N", "B"])
                writer.writerows(workload_rows)
            print(f"  -> wrote workload {workload_path}")

        print(f"  -> wrote {out_path_xlsx}")


if __name__ == "__main__":
    main()


